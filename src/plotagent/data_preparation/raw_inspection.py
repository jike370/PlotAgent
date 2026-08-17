"""Bounded raw-source evidence for Agent-assisted parser selection."""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from plotagent.contracts.data_preparation import DataPreparationRun

MAX_RAW_BYTES = 65_536
MAX_PREVIEW_LINES = 40
MAX_PREVIEW_COLUMNS = 12
MAX_EXCEL_SHEETS = 12


class RawInspectionError(ValueError):
    pass


def _source_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text_evidence(path: Path) -> list[dict[str, Any]]:
    raw = path.read_bytes()[:MAX_RAW_BYTES]
    previews: list[dict[str, Any]] = []
    seen: set[str] = set()
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            text = raw.decode(encoding)
        except UnicodeError:
            continue
        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        if "\x00" in normalized or normalized in seen:
            continue
        seen.add(normalized)
        previews.append(
            {
                "encoding": encoding,
                "lines": normalized.split("\n")[:MAX_PREVIEW_LINES],
                "truncated": (
                    len(raw) == MAX_RAW_BYTES
                    or normalized.count("\n") >= MAX_PREVIEW_LINES
                ),
            }
        )
    return previews


def _excel_evidence(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets[:MAX_EXCEL_SHEETS]:
            rows: list[list[object]] = []
            for values in sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row, MAX_PREVIEW_LINES),
                min_col=1,
                max_col=min(sheet.max_column, MAX_PREVIEW_COLUMNS),
                values_only=True,
            ):
                rows.append(
                    [
                        value if value is None or isinstance(value, (str, int, float, bool))
                        else str(value)
                        for value in values
                    ]
                )
            sheets.append(
                {
                    "sheet": sheet.title,
                    "row_count": sheet.max_row,
                    "column_count": sheet.max_column,
                    "rows": rows,
                    "truncated": (
                        sheet.max_row > MAX_PREVIEW_LINES
                        or sheet.max_column > MAX_PREVIEW_COLUMNS
                    ),
                }
            )
        return sheets
    finally:
        workbook.close()


def inspect_raw_source(path: Path, run: DataPreparationRun) -> dict[str, object]:
    resolved = path.resolve(strict=True)
    if not resolved.is_file() or _source_hash(resolved) != run.source_object_hash:
        raise RawInspectionError("authorized source bytes do not match the preparation run")
    evidence: dict[str, object] = {
        "run_id": run.run_id,
        "source_format": run.probe.source_format,
        "byte_size": run.probe.byte_size,
        "generic_parser_outcome": run.probe.generic_parser_outcome,
        "generic_parser_code": run.probe.generic_parser_code,
        "probed_tables": [table.model_dump(mode="json") for table in run.probe.tables],
    }
    if run.probe.source_format in {"xlsx", "xlsm"}:
        evidence["excel_sheets"] = _excel_evidence(resolved)
    elif run.probe.source_format in {"csv", "tsv", "txt", "dat"}:
        evidence["text_previews"] = _text_evidence(resolved)
    else:
        evidence["raw_preview_unavailable"] = True
    return evidence
