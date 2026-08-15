"""Read-only Excel enumeration and table-region parsing."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from importlib.metadata import version
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from plotagent.contracts.datasets import ExcelSourceCoordinate
from plotagent.importing.errors import ImportErrorCode, ImportProblem
from plotagent.importing.models import (
    ImportRecipe,
    ProvenanceMarker,
    Scalar,
    SourceDatasetArtifact,
    TraceEvent,
)
from plotagent.importing.normalize import (
    build_candidate,
    looks_like_declared_header,
    normalize_excel_scalar,
    stable_hash,
)


@dataclass(frozen=True)
class _SheetMatrix:
    name: str
    rows: tuple[tuple[Scalar, ...], ...]
    provenance: tuple[tuple[int, int, ProvenanceMarker], ...] = ()


@dataclass(frozen=True)
class _Region:
    start_row: int
    end_row: int
    start_col: int
    end_col: int

    @property
    def cell_range(self) -> str:
        return (
            f"{get_column_letter(self.start_col)}{self.start_row}:"
            f"{get_column_letter(self.end_col)}{self.end_row}"
        )


def _read_openpyxl(path: Path) -> tuple[_SheetMatrix, ...]:
    formula_book = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )
    cached_book = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
        keep_vba=False,
    )
    result: list[_SheetMatrix] = []
    try:
        for formula_sheet in formula_book.worksheets:
            cached_sheet = cached_book[formula_sheet.title]
            values: list[tuple[Scalar, ...]] = []
            markers: list[tuple[int, int, ProvenanceMarker]] = []
            formula_rows = formula_sheet.iter_rows()
            cached_rows = cached_sheet.iter_rows()
            for row_number, (formula_row, cached_row) in enumerate(
                zip(formula_rows, cached_rows, strict=False), start=1
            ):
                row_values: list[Scalar] = []
                for column_number, (formula_cell, cached_cell) in enumerate(
                    zip(formula_row, cached_row, strict=False), start=1
                ):
                    coordinate = f"{formula_sheet.title}!{formula_cell.coordinate}"
                    if formula_cell.data_type == "f":
                        cached_value = normalize_excel_scalar(cached_cell.value)
                        row_values.append(cached_value)
                        kind = (
                            "cached_formula_value"
                            if cached_value is not None
                            else "formula_uncached"
                        )
                        markers.append(
                            (
                                row_number,
                                column_number,
                                ProvenanceMarker(kind=kind, coordinate=coordinate),  # type: ignore[arg-type]
                            )
                        )
                        formula_text = str(formula_cell.value or "")
                        folded_formula = formula_text.casefold()
                        if (
                            "[" in formula_text
                            or "http://" in folded_formula
                            or "https://" in folded_formula
                        ):
                            markers.append(
                                (
                                    row_number,
                                    column_number,
                                    ProvenanceMarker(kind="external_link", coordinate=coordinate),
                                )
                            )
                    else:
                        row_values.append(normalize_excel_scalar(formula_cell.value))
                while row_values and row_values[-1] is None:
                    row_values.pop()
                values.append(tuple(row_values))
            while values and not values[-1]:
                values.pop()
            if path.suffix.casefold() == ".xlsm":
                markers.append(
                    (0, 0, ProvenanceMarker(kind="macro_ignored", coordinate="workbook"))
                )
            result.append(
                _SheetMatrix(
                    name=formula_sheet.title,
                    rows=tuple(values),
                    provenance=tuple(markers),
                )
            )
    finally:
        formula_book.close()
        cached_book.close()
    return tuple(result)


def _xlrd_value(book: Any, cell: Any) -> Scalar:
    import xlrd  # type: ignore[import-untyped]

    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        numeric = float(cell.value)
        return int(numeric) if numeric.is_integer() else numeric
    if cell.ctype == xlrd.XL_CELL_DATE:
        return cast(datetime, xlrd.xldate_as_datetime(cell.value, book.datemode))
    return normalize_excel_scalar(cell.value)


def _read_xlrd(path: Path) -> tuple[_SheetMatrix, ...]:
    try:
        import xlrd
    except ImportError as exc:
        raise ImportProblem(
            ImportErrorCode.PARSER_FAILED,
            "缺少 .xls 只读解析组件。",
            "安装受控依赖 xlrd>=2.0.1 后重试。",
        ) from exc
    book = xlrd.open_workbook(path, on_demand=True)
    result: list[_SheetMatrix] = []
    try:
        for sheet in book.sheets():
            rows: list[tuple[Scalar, ...]] = []
            for row_index in range(sheet.nrows):
                values = [
                    _xlrd_value(book, sheet.cell(row_index, col)) for col in range(sheet.ncols)
                ]
                while values and values[-1] is None:
                    values.pop()
                rows.append(tuple(values))
            while rows and not rows[-1]:
                rows.pop()
            result.append(_SheetMatrix(name=str(sheet.name), rows=tuple(rows)))
            sheet.unload()
    finally:
        book.release_resources()
    return tuple(result)


def _nonempty_count(row: tuple[Scalar, ...]) -> int:
    return sum(value is not None for value in row)


def _single_text_note(row: tuple[Scalar, ...]) -> bool:
    values = tuple(value for value in row if value is not None)
    return len(values) == 1 and isinstance(values[0], str)


def _regions(sheet: _SheetMatrix) -> tuple[_Region, ...]:
    groups: list[list[int]] = []
    current: list[int] = []
    for row_number, row in enumerate(sheet.rows, start=1):
        if _nonempty_count(row):
            current.append(row_number)
        elif current:
            groups.append(current)
            current = []
    if current:
        groups.append(current)

    result: list[_Region] = []
    for group in groups:
        while group and _single_text_note(sheet.rows[group[0] - 1]):
            group.pop(0)
        while group and _single_text_note(sheet.rows[group[-1] - 1]):
            group.pop()
        if len(group) < 2:
            continue
        columns = [
            column
            for row_number in group
            for column, value in enumerate(sheet.rows[row_number - 1], start=1)
            if value is not None
        ]
        if not columns:
            continue
        result.append(
            _Region(
                start_row=group[0],
                end_row=group[-1],
                start_col=min(columns),
                end_col=max(columns),
            )
        )
    return tuple(result)


def _row_values(sheet: _SheetMatrix, region: _Region, row_number: int) -> tuple[Scalar, ...]:
    source = sheet.rows[row_number - 1]
    return tuple(
        source[column - 1] if column <= len(source) else None
        for column in range(region.start_col, region.end_col + 1)
    )


def _typed_count(row: tuple[Scalar, ...]) -> int:
    return sum(value is None or not isinstance(value, str) for value in row)


_UNIT_CELL = re.compile(r"^\[([^\[\]]+)\]$")


def _declared_units(row: tuple[Scalar, ...]) -> tuple[str, ...] | None:
    """Return a strict bracketed unit row, never an arbitrary text data row."""

    units: list[str] = []
    found = False
    for value in row:
        if value is None or (isinstance(value, str) and not value.strip()):
            units.append("")
            continue
        if not isinstance(value, str):
            return None
        match = _UNIT_CELL.fullmatch(value.strip())
        if match is None:
            return None
        units.append(match.group(1).strip())
        found = True
    return tuple(units) if found else None


def _header(
    sheet: _SheetMatrix, region: _Region, requested: int | None
) -> tuple[tuple[str, ...], int | None, int]:
    if requested == 0:
        width = region.end_col - region.start_col + 1
        return tuple(f"column_{index}" for index in range(1, width + 1)), None, region.start_row
    first = _row_values(sheet, region, region.start_row)
    second = _row_values(sheet, region, region.start_row + 1)
    if requested is not None:
        if not region.start_row <= requested < region.end_row:
            raise ImportProblem(
                ImportErrorCode.HEADER_AMBIGUOUS,
                "指定的 Excel 表头不在候选区域中。",
                "请选择候选区域内的表头行。",
            )
        header = _row_values(sheet, region, requested)
        return (
            tuple("" if value is None else str(value) for value in header),
            requested,
            requested + 1,
        )
    first_typed = _typed_count(first)
    second_typed = _typed_count(second)
    width = len(first)
    if first_typed == 0 and second_typed > 0:
        return tuple(str(value) for value in first), region.start_row, region.start_row + 1
    if first_typed >= max(1, width // 2):
        return tuple(f"column_{index}" for index in range(1, width + 1)), None, region.start_row
    if first_typed == 0 and second_typed == 0:
        text_first = tuple("" if value is None else str(value) for value in first)
        if looks_like_declared_header(text_first) or _declared_units(second) is not None:
            return text_first, region.start_row, region.start_row + 1
        raise ImportProblem(
            ImportErrorCode.HEADER_AMBIGUOUS,
            f"工作表 {sheet.name} 的前两行都可能是表头。",
            "请选择表头行，或明确该区域无表头。",
            clarification_options=(
                f"line:{region.start_row}",
                f"line:{region.start_row + 1}",
                "none",
            ),
        )
    if second_typed > first_typed:
        return (
            tuple("" if value is None else str(value) for value in first),
            region.start_row,
            region.start_row + 1,
        )
    return tuple(f"column_{index}" for index in range(1, width + 1)), None, region.start_row


def _candidate(
    *,
    path: Path,
    sheet: _SheetMatrix,
    region: _Region,
    source_hash: str,
    parser_name: str,
    parser_version: str,
    header_row: int | None,
) -> SourceDatasetArtifact:
    headers, actual_header, data_start = _header(sheet, region, header_row)
    declared_units = (
        _declared_units(_row_values(sheet, region, data_start))
        if actual_header is not None and data_start <= region.end_row
        else None
    )
    if declared_units is not None:
        data_start += 1
    rows = tuple(_row_values(sheet, region, row) for row in range(data_start, region.end_row + 1))
    coordinates = tuple(
        ExcelSourceCoordinate(
            source_row_id="row:" + stable_hash((source_hash, sheet.name, str(row_number)))[:24],
            workbook_hash=source_hash,
            sheet_name=sheet.name,
            cell_range=(
                f"{get_column_letter(region.start_col)}{row_number}:"
                f"{get_column_letter(region.end_col)}{row_number}"
            ),
        )
        for row_number in range(data_start, region.end_row + 1)
    )
    provenance = tuple(
        marker
        for row, column, marker in sheet.provenance
        if (row == 0 and column == 0)
        or (data_start <= row <= region.end_row and region.start_col <= column <= region.end_col)
    )
    formula_missing = sum(marker.kind == "formula_uncached" for marker in provenance)
    trace = (
        TraceEvent(
            stage="detect",
            code="IMPORT_EXCEL_REGION_DETECTED",
            details={"sheet": sheet.name, "region": region.cell_range},
        ),
        TraceEvent(
            stage="parse",
            code="IMPORT_EXCEL_SHEET_PARSED",
            details={
                "rows": len(rows),
                "columns": region.end_col - region.start_col + 1,
                "formula_uncached": formula_missing,
                "declared_unit_row": declared_units is not None,
            },
        ),
    )
    recipe = ImportRecipe(
        parser_name=parser_name,  # type: ignore[arg-type]
        parser_version=parser_version,
        source_format=path.suffix.casefold().lstrip("."),  # type: ignore[arg-type]
        header_row=actual_header,
        data_start_row=data_start,
        data_end_row=region.end_row,
        workbook=path.name,
        sheet=sheet.name,
        cell_range=region.cell_range,
        column_names=headers,
    )
    return build_candidate(
        display_name=f"{path.stem}:{sheet.name}",
        source_hash=source_hash,
        recipe=recipe,
        headers=headers,
        unit_source_texts=declared_units,
        rows=rows,
        coordinates=coordinates,
        provenance=provenance,
        trace=trace,
    )


def inspect_excel(
    *, path: Path, source_hash: str, selected_sheet: str | None, header_row: int | None
) -> tuple[SourceDatasetArtifact, ...]:
    suffix = path.suffix.casefold()
    if suffix == ".xls":
        sheets = _read_xlrd(path)
        parser_name = "xlrd"
        parser_version = version("xlrd")
    else:
        sheets = _read_openpyxl(path)
        parser_name = "openpyxl"
        parser_version = version("openpyxl")
    if selected_sheet is not None:
        sheets = tuple(sheet for sheet in sheets if sheet.name == selected_sheet)
        if not sheets:
            raise ImportProblem(
                ImportErrorCode.NO_DATA,
                "指定的工作表不存在。",
                "请从只读枚举的工作表列表中选择。",
            )

    candidates: list[SourceDatasetArtifact] = []
    for sheet in sheets:
        regions = _regions(sheet)
        if len(regions) > 1:
            raise ImportProblem(
                ImportErrorCode.REGION_AMBIGUOUS,
                f"工作表 {sheet.name} 包含多个同等合理的数据区域。",
                "请选择一个单元格区域后继续导入。",
                clarification_options=tuple(region.cell_range for region in regions),
            )
        if not regions:
            continue
        candidates.append(
            _candidate(
                path=path,
                sheet=sheet,
                region=regions[0],
                source_hash=source_hash,
                parser_name=parser_name,
                parser_version=parser_version,
                header_row=header_row,
            )
        )
    if not candidates:
        raise ImportProblem(
            ImportErrorCode.NO_DATA,
            "工作簿中没有找到可导入的二维数据区域。",
            "请确认至少一个工作表包含表头和数据行。",
        )
    return tuple(candidates)
