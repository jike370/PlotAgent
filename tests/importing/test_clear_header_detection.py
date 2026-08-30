from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

import plotagent.importing.text as text_module
from plotagent.importing import Clarification, Imported, inspect_source


def test_clear_categorical_csv_header_does_not_require_confirmation(tmp_path: Path) -> None:
    source = tmp_path / "confusion.csv"
    source.write_text("Actual,Predicted\nCat,Cat\nDog,Cat\n", encoding="utf-8")

    result = inspect_source(source)

    assert isinstance(result, Imported)
    assert tuple(field.name for field in result.sources[0].source_dataset.field_schema) == (
        "Actual",
        "Predicted",
    )


def test_clear_categorical_excel_sheet_does_not_block_other_sheets(tmp_path: Path) -> None:
    source = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Numeric"
    first.append(("X", "Y"))
    first.append((1, 2))
    second = workbook.create_sheet("S61_raw")
    second.append(("Actual", "Predicted"))
    second.append(("Cat", "Cat"))
    second.append(("Dog", "Cat"))
    workbook.save(source)

    result = inspect_source(source)

    assert isinstance(result, Imported)
    assert tuple(item.recipe.sheet for item in result.sources) == ("Numeric", "S61_raw")


def test_excel_header_clarification_is_scoped_to_one_sheet(tmp_path: Path) -> None:
    source = tmp_path / "scoped-headers.xlsx"
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(("category", "value"))
    data.append((0, 0.33852964))
    data.append((1, 0.30398383))
    lineage = workbook.create_sheet("Lineage")
    lineage.append(("Data lineage", "Value"))
    lineage.append(("Paper", "Example"))
    lineage.append(("Source", "Figure 2"))
    workbook.save(source)

    clarification = inspect_source(source)
    assert isinstance(clarification, Clarification)
    assert tuple(option.value for option in clarification.options) == (
        "sheet:Lineage|line:1",
        "sheet:Lineage|line:2",
        "sheet:Lineage|none",
    )

    result = inspect_source(source, header_rows={"Lineage": 2})
    assert isinstance(result, Imported)
    data_artifact = next(item for item in result.sources if item.recipe.sheet == "Data")
    assert tuple(field.name for field in data_artifact.source_dataset.field_schema) == (
        "category",
        "value",
    )
    assert data_artifact.rows == ((0, 0.33852964), (1, 0.30398383))


def test_bracketed_excel_unit_row_is_metadata_not_sample_data(tmp_path: Path) -> None:
    source = tmp_path / "units.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(("X", "Response_mV"))
    sheet.append(("[mm]", "[mV]"))
    sheet.append((1, 4.28))
    sheet.append((2, 5.31))
    workbook.save(source)

    result = inspect_source(source)

    assert isinstance(result, Imported)
    artifact = result.sources[0]
    assert artifact.rows == ((1, 4.28), (2, 5.31))
    assert artifact.recipe.data_start_row == 3
    assert tuple(field.name for field in artifact.source_dataset.field_schema) == (
        "X",
        "Response_mV",
    )
    assert tuple(field.logical_type for field in artifact.source_dataset.field_schema) == (
        "numeric",
        "numeric",
    )
    assert tuple(field.unit.source_text for field in artifact.source_dataset.field_schema) == (
        "mm",
        "mV",
    )


def test_explicit_scientific_header_suffixes_are_imported_as_unit_suggestions(
    tmp_path: Path,
) -> None:
    source = tmp_path / "signals.csv"
    source.write_text("Time_s,Signal_mV\n0,1.5\n1,2.5\n", encoding="utf-8")

    result = inspect_source(source)

    assert isinstance(result, Imported)
    fields = result.sources[0].source_dataset.field_schema
    assert tuple(field.name for field in fields) == ("Time_s", "Signal_mV")
    assert tuple(field.unit.source_text for field in fields) == ("s", "mV")
    assert tuple(field.unit.kind for field in fields) == ("opaque", "opaque")


def test_arbitrary_header_suffix_is_not_invented_as_a_unit(tmp_path: Path) -> None:
    source = tmp_path / "identities.csv"
    source.write_text("sample_id,value\nA,1\nB,2\n", encoding="utf-8")

    result = inspect_source(source)

    assert isinstance(result, Imported)
    fields = result.sources[0].source_dataset.field_schema
    assert tuple(field.unit.kind for field in fields) == ("dimensionless", "dimensionless")


def test_instrument_table_header_after_preamble_is_detected_and_trailing_empty_field_dropped(
    tmp_path: Path,
) -> None:
    source = tmp_path / "instrument.txt"
    source.write_text(
        "Instrument=Analyzer 7\n"
        "Operator=Lab A\n"
        "\n"
        "     Angle,       PSD,\n"
        "0.100,12.5,\n"
        "0.200,13.75,\n"
        "0.300,15.0,\n",
        encoding="utf-8",
    )

    result = inspect_source(source)

    assert isinstance(result, Imported)
    artifact = result.sources[0]
    assert artifact.recipe.header_row == 4
    assert artifact.recipe.data_start_row == 5
    assert artifact.recipe.column_names == ("Angle", "PSD")
    assert artifact.rows == ((0.1, 12.5), (0.2, 13.75), (0.3, 15.0))
    assert tuple(field.logical_type for field in artifact.source_dataset.field_schema) == (
        "numeric",
        "numeric",
    )
    assert artifact.coordinates[0].line_start == 5
    assert artifact.instrument_metadata == {"Instrument": "Analyzer 7", "Operator": "Lab A"}
    parsed = next(event for event in artifact.trace if event.code == "IMPORT_TEXT_BLOCK_PARSED")
    assert parsed.details == {"rows": 3, "columns": 2, "discarded_empty_columns": 1}


def test_large_text_preamble_boundary_is_computed_once(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "large.csv"
    source.write_text(
        "x,y\n" + "\n".join(f"{index},{index * 2}" for index in range(2000)) + "\n",
        encoding="utf-8",
    )
    builtin_min = min
    calls = 0

    def counting_min(*args: Any, **kwargs: Any) -> Any:
        nonlocal calls
        calls += 1
        return builtin_min(*args, **kwargs)

    monkeypatch.setattr(text_module, "min", counting_min, raising=False)

    result = inspect_source(source)

    assert isinstance(result, Imported)
    assert result.sources[0].source_dataset.data_ref.row_count == 2000
    assert calls == 1


def test_sparse_real_field_is_not_dropped_as_trailing_empty_serialization(tmp_path: Path) -> None:
    source = tmp_path / "sparse.csv"
    source.write_text("X,Y,Note,\n1,2,,\n2,3,keep,\n", encoding="utf-8")

    result = inspect_source(source)

    assert isinstance(result, Imported)
    artifact = result.sources[0]
    assert artifact.recipe.column_names == ("X", "Y", "Note")
    assert artifact.rows == ((1, 2, None), (2, 3, "keep"))


def test_unit_registry_preserves_scientific_prefix_case(tmp_path: Path) -> None:
    source = tmp_path / "resistance.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("Value (MΩ)", "Signal (mΩ)"))
    sheet.append((2, 3))
    sheet.append((4, 5))
    workbook.save(source)

    result = inspect_source(source)

    assert isinstance(result, Imported)
    fields = result.sources[0].source_dataset.field_schema
    assert tuple(field.unit.source_text for field in fields) == ("MΩ", "mΩ")
    assert tuple(field.unit.canonical_unit for field in fields) == ("Mohm", "mohm")


def test_ambiguous_single_letter_suffixes_are_not_asserted_as_units(tmp_path: Path) -> None:
    source = tmp_path / "ambiguous_suffixes.csv"
    source.write_text(
        "distance_m,mass_g,duration_h,pressure_kPa\n1,2,3,4\n",
        encoding="utf-8",
    )

    result = inspect_source(source)

    assert isinstance(result, Imported)
    fields = result.sources[0].source_dataset.field_schema
    assert tuple(field.unit.source_text for field in fields) == ("", "", "", "kPa")
    assert tuple(field.unit.kind for field in fields) == (
        "dimensionless",
        "dimensionless",
        "dimensionless",
        "opaque",
    )
